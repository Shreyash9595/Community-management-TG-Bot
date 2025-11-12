# main_webwebhook.py
import os, math, datetime, pandas as pd
from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters, CallbackQueryHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# --- NEW IMPORTS ---
from flask import Flask, request
import asyncio
import threading
import concurrent.futures

# --- CONFIGURATION ---
TOKEN = os.environ.get('TELEGRAM_TOKEN')
if not TOKEN:
    raise ValueError("No TELEGRAM_TOKEN set in environment variables")

DOWNLOAD_DIR = "downloads"
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

hunt_data = None
filtered_players = []
pending_selections = {}
original_hunt_data = None  # Store original data for include/exclude operations

# ---------- UPDATED Date filtering function ----------
def filter_players_by_guild_presence(df):
    global filtered_players
    filtered_players = []
    
    try:
        filtered_df = df.copy()
        
        # Remove "Total" row from Name column first
        original_count_before_total = len(filtered_df)
        filtered_df = filtered_df[filtered_df['Name'].str.lower() != 'total']
        total_removed = original_count_before_total - len(filtered_df)
        if total_removed > 0:
            print(f"Removed {total_removed} 'Total' row(s)")
        
        # Check if we have the required date columns
        date_cols_present = all(col in filtered_df.columns for col in ["First Hunt Time", "Last Hunt Time"])
        
        if not date_cols_present:
            print("Date columns not found - skipping date-based filtering")
            return filtered_df, []
        
        # Convert date columns to datetime
        filtered_df['First Hunt Time'] = pd.to_datetime(filtered_df['First Hunt Time'], errors='coerce')
        filtered_df['Last Hunt Time'] = pd.to_datetime(filtered_df['Last Hunt Time'], errors='coerce')
        
        # Identify players with valid dates for filtering
        valid_dates_mask = filtered_df['First Hunt Time'].notna() & filtered_df['Last Hunt Time'].notna()
        players_with_valid_dates = filtered_df[valid_dates_mask].copy()
        players_without_dates = filtered_df[~valid_dates_mask].copy()
        
        print(f"Players with valid dates: {len(players_with_valid_dates)}")
        print(f"Players without valid dates (0% hunters): {len(players_without_dates)}")
        
        # Calculate days in guild only for players with valid dates
        players_with_valid_dates['Days In Guild'] = (players_with_valid_dates['Last Hunt Time'] - players_with_valid_dates['First Hunt Time']).dt.days.abs()
        
        # Filter out players with less than 4 days in guild (only from those with valid dates)
        players_to_filter = players_with_valid_dates[players_with_valid_dates['Days In Guild'] < 4]
        
        # Also include players without dates (0% hunters) in the filtered list
        players_without_dates['Days In Guild'] = 0  # Mark as 0 days for tracking
        players_without_dates['Reason'] = '0% hunter (no date data)'
        
        # Combine both filtered groups
        all_filtered_players = pd.concat([players_to_filter, players_without_dates], ignore_index=True)
        
        # Prepare filtered players info for display
        filtered_players_info = []
        for _, player in all_filtered_players.iterrows():
            player_info = {
                'Name': player['Name'],
                'Days In Guild': player.get('Days In Guild', 0),
                'Reason': player.get('Reason', 'Less than 4 days in guild')
            }
            filtered_players_info.append(player_info)
        
        # Keep ONLY players with 4+ days in guild (exclude both <4 days AND 0% hunters)
        players_with_enough_days = players_with_valid_dates[players_with_valid_dates['Days In Guild'] >= 4]
        
        # Final dataframe contains ONLY players with 4+ days in guild
        final_df = players_with_enough_days.copy()
        
        # Drop the temporary column
        if 'Days In Guild' in final_df.columns:
            final_df = final_df.drop('Days In Guild', axis=1)
        
        print(f"Original count: {original_count_before_total}")
        print(f"After removing 'Total': {len(filtered_df)}") 
        print(f"Filtered out {len(players_to_filter)} players with less than 4 days in guild")
        print(f"Filtered out {len(players_without_dates)} 0% hunters (no date data)")
        print(f"Final count (only players with 4+ days in guild): {len(final_df)}")
        
        return final_df, filtered_players_info
        
    except Exception as e:
        print(f"Error in date filtering: {e}")
        return df, []

# ---------- Helper functions ----------
def parse_percent_cell(raw):
    if raw is None: return None
    if isinstance(raw, (int, float)): return float(raw)
    s = str(raw).strip().replace(",", "").replace(" ", "")
    if not s or s.lower() in ("-", "--", "n/a", "na"): return None
    if s.endswith("%"): s = s[:-1]
    try: return float(s)
    except: return None

def convert_and_round_series(series):
    parsed = series.map(parse_percent_cell)
    numeric = parsed.dropna().astype(float)
    if numeric.empty: 
        return pd.Series(0, index=series.index)
    
    if numeric.max() <= 100:
        numeric *= 100
    
    rounded = numeric.round().astype(int)
    
    def get_value(i):
        if not pd.isna(parsed[i]):
            return int(rounded.get(i, 0))
        else:
            return 0
    
    return parsed.index.to_series().map(get_value)

def format_and_color_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    widths = {1: 30, 2: 18, 3: 18, 4: 40} 
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 18)

    # Color coding for top 3 ranks
    rank_colors = {
        1: "e79d5e",  # First rank color
        2: "5e95e7",  # Second rank color  
        3: "faa3d1"   # Third rank color
    }
    
    # Find Kidnape me row
    kidnape_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value and "kidnape me" in str(ws.cell(r, 1).value).lower():
            kidnape_row = r
            break
    
    # Add rank numbers to issues column for top 3
    for r in range(2, min(5, ws.max_row + 1)):
        rank = r - 1
        issue_cell = ws.cell(r, 4)
        if rank == 1:
            issue_cell.value = "Rank 1"
        elif rank == 2:
            issue_cell.value = "Rank 2" 
        elif rank == 3:
            issue_cell.value = "Rank 3"
        # Keep the text but no background color
        issue_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Apply colors to top 3 rows
    for r in range(2, min(5, ws.max_row + 1)):
        rank = r - 1
        if rank in rank_colors:
            # Set the color for the row (only first 3 columns)
            fill = PatternFill(start_color=rank_colors[rank], end_color=rank_colors[rank], fill_type="solid")
            
            # Apply color to first 3 columns only
            for c in range(1, 4):
                cell = ws.cell(r, c)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")
            
            # Issues column - no color, left aligned (already set above)
    
    # Handle Kidnape me row
    if kidnape_row:
        # Make Kidnape me row white (first 3 columns)
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for c in range(1, 4):
            kidnape_cell = ws.cell(kidnape_row, c)
            kidnape_cell.fill = white_fill
            kidnape_cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")
        
        # Add BANK text to issues column for Kidnape me row (no color, left aligned)
        bank_issue_cell = ws.cell(kidnape_row, 4)
        bank_issue_cell.value = "BANK"
        bank_issue_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Format the rest of the rows (after top 3)
    start_row = 5
    if kidnape_row and kidnape_row < 5:  # Adjust if Kidnape me is in top rows
        start_row = 6
    
    for r in range(start_row, ws.max_row + 1):
        # Skip Kidnape me row (already formatted)
        if kidnape_row and r == kidnape_row:
            continue
            
        try:
            # Extract numeric value from percentage string (e.g., "150%" -> 150)
            cell_value = ws.cell(r, 3).value
            if cell_value and isinstance(cell_value, str) and cell_value.endswith('%'):
                p = float(cell_value.rstrip('%'))
            else:
                p = float(cell_value or 0)
        except:
            p = 0
        
        if p >= 100: color = "90EE90"
        elif p >= 90: color = "F5DEB3"
        elif p >= 70: color = "FFFF99"
        else: color = "FFB3B3"
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for c in range(1, 4):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")

        # Issues column - no color, left aligned
        issue_cell = ws.cell(r, 4)
        issue_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Remove "Issues" header and replace with empty string
    ws.cell(1, 4).value = ""

    wb.save(path)

# --- FIXED: Background processing function ---
def process_file_in_background(local_path, chat_id, message_id):
    """Process file in background and send result when done"""
    try:
        # Use a new event loop for the background thread
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        # Create a new bot instance for this background task
        bot = Bot(token=TOKEN)
        
        df = pd.read_excel(local_path)
        cols = ["Name","Points (Hunt)","Goal Percentage (Hunt)"]
        if not all(c in df.columns for c in cols):
            loop.run_until_complete(bot.edit_message_text(
                chat_id=chat_id, 
                message_id=message_id,
                text="❌ Missing required columns."
            ))
            return
        
        # Store original data for include/exclude operations
        global original_hunt_data
        original_hunt_data = df.copy()
        
        # Apply filtering (this now keeps ONLY players with 4+ days in guild)
        global filtered_players
        df, date_filtered_players = filter_players_by_guild_presence(df)
        filtered_players = date_filtered_players  # Store filtered players
        
        # Count removed "Total" rows separately
        original_df = pd.read_excel(local_path)
        total_removed_count = len(original_df[original_df['Name'].str.lower() == 'total'])
        
        # Prepare filter message
        filter_message_parts = []
        if total_removed_count > 0:
            filter_message_parts.append(f"📊 Removed {total_removed_count} 'Total' row(s)")
        
        # Count different types of filtered players
        less_than_4_days = len([p for p in filtered_players if p.get('Reason') == 'Less than 4 days in guild'])
        zero_percent_hunters = len([p for p in filtered_players if p.get('Reason') == '0% hunter (no date data)'])
        
        if less_than_4_days > 0:
            filter_message_parts.append(f"📅 Filtered out {less_than_4_days} players with less than 4 days in guild")
        if zero_percent_hunters > 0:
            filter_message_parts.append(f"🎯 Filtered out {zero_percent_hunters} 0% hunters (no date data)")
        
        filter_message = "\n" + "\n".join(filter_message_parts) if filter_message_parts else ""
        
        df = df[cols].copy()
        df["Issues"] = "" 
        df["Points (Hunt)"] = pd.to_numeric(df["Points (Hunt)"], errors="coerce").fillna(0).astype(int)
        
        # Convert goal percentage and add % sign
        goal_percentages = convert_and_round_series(df["Goal Percentage (Hunt)"])
        df["Goal_Percentage_Numeric"] = goal_percentages.clip(lower=0)
        df["Goal Percentage (Hunt)"] = df["Goal_Percentage_Numeric"].astype(str) + "%"
        
        # Sort by the numeric column
        df = df.sort_values("Goal_Percentage_Numeric", ascending=False).reset_index(drop=True)
        df = df.drop("Goal_Percentage_Numeric", axis=1)  # Remove temporary column
        
        global hunt_data
        hunt_data = df.copy()
        
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(DOWNLOAD_DIR, f"Processed_Report_{ts}.xlsx")
        
        df.to_excel(out_path, index=False)
        format_and_color_excel(out_path)
        
        # Update the message and send document
        loop.run_until_complete(bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=f"✅ File processed!{filter_message}\nUse /problem, /generate, /i, /summary, /top10, /bottom10, /bottom20, /0hunt, /clear, /filteredplayers, /include, /exclude"
        ))
        
        # Send the document
        with open(out_path, 'rb') as file:
            loop.run_until_complete(bot.send_document(
                chat_id=chat_id,
                document=file,
                filename=os.path.basename(out_path)
            ))
            
    except Exception as e:
        error_msg = f"❌ Error processing file: {str(e)}"
        try:
            bot = Bot(token=TOKEN)
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=error_msg
            ))
        except:
            print(f"Failed to send error message: {error_msg}")

# --- UPDATED handle_file ---
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc or not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("❌ Please upload a .xlsx file.")
        return
        
    local_path = os.path.join(DOWNLOAD_DIR, doc.file_name)
    msg = await update.message.reply_text("⏳ Processing file... This may take a moment.")
    
    # Download file first
    await (await doc.get_file()).download_to_drive(local_path)
    
    # Start background processing
    thread = threading.Thread(
        target=process_file_in_background,
        args=(local_path, msg.chat_id, msg.message_id)
    )
    thread.daemon = True
    thread.start()

# --- UPDATED Button handler for player selection ---
async def handle_player_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    data = query.data
    user_id = query.from_user.id
    
    if data.startswith("select_player_"):
        # Format: select_player_{index}_{user_id}
        parts = data.split("_")
        if len(parts) >= 4:
            player_index = int(parts[2])
            target_user_id = int(parts[3])
            
            # Check if this callback is for the right user
            if user_id != target_user_id:
                await query.edit_message_text("This selection is not for you.")
                return
            
            if user_id in pending_selections and player_index < len(pending_selections[user_id]['players']):
                selection_data = pending_selections[user_id]
                player_name = selection_data['players'][player_index]
                issue_text = selection_data['issue_text']
                
                # Find the player in hunt_data
                results = hunt_data[hunt_data['Name'].str.contains(player_name, case=False, na=False)]
                if len(results) == 1:
                    player_index_df = results.index[0]
                    hunt_data.loc[player_index_df, 'Issues'] = issue_text
                    
                    await query.edit_message_text(
                        f"✅ Updated issue for *{player_name}*:\n"
                        f"_{issue_text}_", 
                        parse_mode="Markdown"
                    )
                    
                    # Clean up pending selection
                    del pending_selections[user_id]
                else:
                    await query.edit_message_text(f"❌ Could not find player: {player_name}")
    
    elif data.startswith("exclude_player_"):
        # Format: exclude_player_{index}_{user_id}
        parts = data.split("_")
        if len(parts) >= 4:
            player_index = int(parts[2])
            target_user_id = int(parts[3])
            
            # Check if this callback is for the right user
            if user_id != target_user_id:
                await query.edit_message_text("This selection is not for you.")
                return
            
            if user_id in pending_selections and player_index < len(pending_selections[user_id]['players']):
                selection_data = pending_selections[user_id]
                player_name = selection_data['players'][player_index]
                
                # Find the player in hunt_data
                results = hunt_data[hunt_data['Name'].str.contains(player_name, case=False, na=False)]
                if len(results) == 1:
                    player_to_remove = results.iloc[0]
                    
                    # Add player to filtered_players with reason
                    filtered_player_info = {
                        'Name': player_to_remove['Name'],
                        'Days In Guild': 'Manually excluded',
                        'Reason': 'Manually excluded by admin'
                    }
                    filtered_players.append(filtered_player_info)
                    
                    # Remove from hunt_data
                    hunt_data = hunt_data[hunt_data['Name'] != player_to_remove['Name']]
                    
                    await query.edit_message_text(
                        f"✅ Removed *{player_name}* from the report and added to filtered players.\n"
                        f"Use `/generate` to create updated report.", 
                        parse_mode="Markdown"
                    )
                    
                    # Clean up pending selection
                    del pending_selections[user_id]
                else:
                    await query.edit_message_text(f"❌ Could not find player: {player_name}")
    
    elif data.startswith("cancel_selection_"):
        target_user_id = int(data.split("_")[2])
        if user_id == target_user_id and user_id in pending_selections:
            del pending_selections[user_id]
            await query.edit_message_text("❌ Selection cancelled.")
    
    elif data.startswith("cancel_exclude_"):
        target_user_id = int(data.split("_")[2])
        if user_id == target_user_id and user_id in pending_selections:
            del pending_selections[user_id]
            await query.edit_message_text("❌ Exclusion cancelled.")

# --- UPDATED problem command with buttons ---
async def problem(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global hunt_data
    if hunt_data is None: 
        return await update.message.reply_text("⚠️ Upload a file first.")
    
    if not context.args:
        return await update.message.reply_text("Please provide a name and issue. \nExample: `/problem Player Name is busy`", parse_mode="Markdown")
    
    args = context.args
    user_id = update.effective_user.id
    
    # Try to find exact match first
    found_match = False
    
    for i in range(len(args), 0, -1):
        search_name = " ".join(args[:i])
        issue_text = " ".join(args[i:])
        
        results = hunt_data[hunt_data['Name'].str.contains(search_name, case=False, na=False)]
        
        if len(results) == 1:
            player_index = results.index[0]
            player_name = results.loc[player_index, 'Name']

            if not issue_text:
                return await update.message.reply_text(f"Found *{player_name}*. Please provide an issue description.", parse_mode="Markdown")
            
            hunt_data.loc[player_index, 'Issues'] = issue_text
            
            await update.message.reply_text(f"✅ Updated issue for *{player_name}*:\n"
                                              f"_{issue_text}_", parse_mode="Markdown")
            found_match = True
            break
        
        elif len(results) > 1:
            # Multiple matches found - show buttons
            if not issue_text:
                return await update.message.reply_text(f"Found *{len(results)}* matches for *{search_name}*. Please provide an issue description after the name.", parse_mode="Markdown")
            
            # Store the pending selection
            pending_selections[user_id] = {
                'players': results['Name'].tolist(),
                'issue_text': issue_text
            }
            
            # Create buttons for selection
            keyboard = []
            for idx, player_name in enumerate(results['Name'].head(10)):  # Limit to 10 players
                keyboard.append([InlineKeyboardButton(player_name, callback_data=f"select_player_{idx}_{user_id}")])
            
            keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data=f"cancel_selection_{user_id}")])
            
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"Found *{len(results)}* matches. Please select a player:",
                reply_markup=reply_markup,
                parse_mode="Markdown"
            )
            found_match = True
            break

    if not found_match:
        return await update.message.reply_text(f"❌ Player not found matching: *{' '.join(args)}*", parse_mode="Markdown")

# --- UPDATED show_filtered_players command ---
async def show_filtered_players(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not filtered_players:
        await update.message.reply_text("ℹ️ No players were filtered out in the last processing.")
        return
    
    try:
        txt = ["🚫 *Filtered Players*", "━━━━━━━━━━━━━━━"]
        
        for i, player in enumerate(filtered_players, 1):
            name = player.get('Name', 'Unknown')
            days = player.get('Days In Guild', 'Unknown')
            reason = player.get('Reason', 'Less than 4 days in guild')
            
            if reason == 'Manually excluded by admin':
                txt.append(f"{i}. *{name}* — {reason}")
            elif reason == '0% hunter (no date data)':
                txt.append(f"{i}. *{name}* — {reason}")
            else:
                txt.append(f"{i}. *{name}* — {days} days")
        
        message_text = "\n".join(txt)
        
        # If still too long, split into multiple messages
        if len(message_text) > 4000:
            parts = []
            current_part = ["🚫 *Filtered Players*", "━━━━━━━━━━━━━━━"]
            
            for i, player in enumerate(filtered_players, 1):
                name = player.get('Name', 'Unknown')
                days = player.get('Days In Guild', 'Unknown')
                reason = player.get('Reason', 'Less than 4 days in guild')
                
                if reason == 'Manually excluded by admin':
                    line = f"{i}. *{name}* — {reason}"
                elif reason == '0% hunter (no date data)':
                    line = f"{i}. *{name}* — {reason}"
                else:
                    line = f"{i}. *{name}* — {days} days"
                
                if len("\n".join(current_part + [line])) > 4000:
                    parts.append("\n".join(current_part))
                    current_part = [line]
                else:
                    current_part.append(line)
            
            if current_part:
                parts.append("\n".join(current_part))
            
            for part in parts:
                await update.message.reply_text(part, parse_mode="Markdown")
        else:
            await update.message.reply_text(message_text, parse_mode="Markdown")
            
    except Exception as e:
        await update.message.reply_text(f"❌ Error showing filtered players: {e}")

# --- NEW: Include command ---
async def include_player(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global hunt_data, filtered_players, original_hunt_data
    
    if hunt_data is None or original_hunt_data is None:
        return await update.message.reply_text("⚠️ Upload a file first.")
    
    if not context.args:
        return await update.message.reply_text("Please provide a player name to include. \nExample: `/include Player Name`", parse_mode="Markdown")
    
    player_name = " ".join(context.args)
    
    # Search in filtered players
    found_in_filtered = None
    for player in filtered_players:
        if player_name.lower() in player.get('Name', '').lower():
            found_in_filtered = player
            break
    
    if not found_in_filtered:
        return await update.message.reply_text(f"❌ Player *{player_name}* not found in filtered players list. Use `/filteredplayers` to see available players.", parse_mode="Markdown")
    
    # Search in original data to get full player info
    original_player = original_hunt_data[original_hunt_data['Name'].str.contains(player_name, case=False, na=False)]
    
    if original_player.empty:
        return await update.message.reply_text(f"❌ Could not find original data for *{player_name}*.", parse_mode="Markdown")
    
    # Add player to hunt_data
    player_to_add = original_player.iloc[0]
    
    # Create a new row with the required columns
    new_row = {
        'Name': player_to_add['Name'],
        'Points (Hunt)': player_to_add.get('Points (Hunt)', 0),
        'Goal Percentage (Hunt)': player_to_add.get('Goal Percentage (Hunt)', '0%'),
        'Issues': ''
    }
    
    # Convert to DataFrame and concatenate
    new_row_df = pd.DataFrame([new_row])
    hunt_data = pd.concat([hunt_data, new_row_df], ignore_index=True)
    
    # Remove from filtered_players
    filtered_players = [p for p in filtered_players if player_name.lower() not in p.get('Name', '').lower()]
    
    await update.message.reply_text(f"✅ Added *{player_name}* back to the report. Use `/generate` to create updated report.", parse_mode="Markdown")

# --- IMPROVED: Exclude command with better name matching ---
async def exclude_player(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global hunt_data, filtered_players
    
    if hunt_data is None:
        return await update.message.reply_text("⚠️ Upload a file first.")
    
    if not context.args:
        return await update.message.reply_text("Please provide a player name to exclude. \nExample: `/exclude Player Name`", parse_mode="Markdown")
    
    args = context.args
    user_id = update.effective_user.id
    
    # Try to find exact match first (same logic as /i and /problem commands)
    found_match = False
    
    for i in range(len(args), 0, -1):
        search_name = " ".join(args[:i])
        remaining_args = " ".join(args[i:])
        
        # Search in current hunt_data
        results = hunt_data[hunt_data['Name'].str.contains(search_name, case=False, na=False)]
        
        if len(results) == 1:
            player_name = results.iloc[0]['Name']
            
            # Remove the player from hunt_data and add to filtered_players
            player_to_remove = results.iloc[0]
            
            # Add player to filtered_players with reason
            filtered_player_info = {
                'Name': player_to_remove['Name'],
                'Days In Guild': 'Manually excluded',
                'Reason': 'Manually excluded by admin'
            }
            filtered_players.append(filtered_player_info)
            
            # Remove from hunt_data
            hunt_data = hunt_data[hunt_data['Name'] != player_to_remove['Name']]
            
            await update.message.reply_text(
                f"✅ Removed *{player_name}* from the report and added to filtered players.\n"
                f"Use `/generate` to create updated report.", 
                parse_mode="Markdown"
            )
            found_match = True
            break
        
        elif len(results) > 1:
            # Multiple matches found - show buttons for selection
            pending_selections[user_id] = {
                'players': results['Name'].tolist(),
                'action': 'exclude'  # Track that this is for exclusion
            }
            
            # Create buttons for selection
            keyboard = []
            for idx, player_name in enumerate(results['Name'].head(10)):  # Limit to 10 players
                keyboard.append([InlineKeyboardButton(player_name, callback_data=f"exclude_player_{idx}_{user_id}")])
            
            keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data=f"cancel_exclude_{user_id}")])
            
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"Found *{len(results)}* matches for '*{search_name}*'. Please select a player to exclude:",
                reply_markup=reply_markup,
                parse_mode="Markdown"
            )
            found_match = True
            break

    if not found_match:
        # If no matches found, show suggestions
        search_name = " ".join(args)
        all_players = hunt_data['Name'].tolist()
        
        # Find similar names (case-insensitive partial matches)
        similar_players = [name for name in all_players if search_name.lower() in name.lower()]
        
        if similar_players:
            player_list = "\n".join([f"- {name}" for name in similar_players[:5]])  # Show top 5 matches
            await update.message.reply_text(
                f"❌ No exact match found for '*{search_name}*'. Similar names:\n{player_list}",
                parse_mode="Markdown"
            )
        else:
            await update.message.reply_text(
                f"❌ Player not found matching: *{search_name}*\n"
                f"Use `/i PlayerName` to search for available players.",
                parse_mode="Markdown"
            )

# --- Rest of the command handlers ---
async def top10(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data
    if hunt_data is None: return await update.message.reply_text("⚠️ Upload a file first.")
    
    # Create a temporary DataFrame with numeric goal percentage for sorting
    temp_df = hunt_data.copy()
    temp_df['Goal_Numeric'] = temp_df['Goal Percentage (Hunt)'].str.rstrip('%').astype(float)
    top = temp_df.nlargest(10, 'Goal_Numeric')
    
    medals = ["🥇","🥈","🥉"]+[f"{i+1}️⃣" for i in range(3,10)]
    txt = ["🏆 *Top 10 Hunters*","━━━━━━━━━━━━━━━"]+[f"{medals[i]} {row['Name']} — {row['Goal Percentage (Hunt)']} ({row['Points (Hunt)']} pts)" for i, (_, row) in enumerate(top.iterrows())]
    await update.message.reply_text("\n".join(txt), parse_mode="Markdown")

async def bottom10(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data
    if hunt_data is None: return await update.message.reply_text("⚠️ Upload a file first.")
    
    # Create a temporary DataFrame with numeric goal percentage for sorting
    temp_df = hunt_data.copy()
    temp_df['Goal_Numeric'] = temp_df['Goal Percentage (Hunt)'].str.rstrip('%').astype(float)
    bottom = temp_df.nsmallest(10, 'Goal_Numeric')
    
    txt = ["🔻 *Bottom 10 Hunters*","━━━━━━━━━━━━━━━"]+[f"{i}. {row['Name']} — {row['Goal Percentage (Hunt)']} ({row['Points (Hunt)']} pts)" for i, (_, row) in enumerate(bottom.iterrows(), 1)]
    await update.message.reply_text("\n".join(txt), parse_mode="Markdown")

async def bottom20(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data
    if hunt_data is None: return await update.message.reply_text("⚠️ Upload a file first.")
    
    # Create a temporary DataFrame with numeric goal percentage for sorting
    temp_df = hunt_data.copy()
    temp_df['Goal_Numeric'] = temp_df['Goal Percentage (Hunt)'].str.rstrip('%').astype(float)
    bottom = temp_df.nsmallest(20, 'Goal_Numeric')
    
    txt = ["🔻 *Bottom 20 Hunters*","━━━━━━━━━━━━━━━"]+[f"{i}. {row['Name']} — {row['Goal Percentage (Hunt)']} ({row['Points (Hunt)']} pts)" for i, (_, row) in enumerate(bottom.iterrows(), 1)]
    await update.message.reply_text("\n".join(txt), parse_mode="Markdown")

async def summary(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data
    if hunt_data is None: 
        return await update.message.reply_text("⚠️ Upload a file first.")
    try:
        total_hunters = len(hunt_data)
        total_points = hunt_data["Points (Hunt)"].sum()
        
        # Calculate average goal percentage (convert from string with % sign)
        goal_numeric = hunt_data['Goal Percentage (Hunt)'].str.rstrip('%').astype(float)
        avg_goal = goal_numeric.mean()
        
        # Get top and lowest hunters
        top_idx = goal_numeric.idxmax()
        bottom_idx = goal_numeric.idxmin()
        top_hunter = hunt_data.loc[top_idx]
        lowest_hunter = hunt_data.loc[bottom_idx]
        
        remark = "🔥 Excellent performance this week!"
        if avg_goal < 50: remark = "💪 Let's push for more next week!"
        txt = [
            "📊 *Weekly Hunt Summary*", "━━━━━━━━━━━━━━━",
            f"👥 *Total Hunters:* {total_hunters}",
            f"⚔️ *Total Hunt Points:* {total_points:,}",
            f"🎯 *Avg Goal % (Hunt):* {avg_goal:.1f}%", "━━━━━━━━━━━━━━━",
            f"🏆 *Top Hunter:* {top_hunter['Name']} — {top_hunter['Goal Percentage (Hunt)']}",
            f"💀 *Lowest Hunter:* {lowest_hunter['Name']} — {lowest_hunter['Goal Percentage (Hunt)']}",
            "━━━━━━━━━━━━━━━", remark
        ]
        await update.message.reply_text("\n".join(txt), parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Error creating summary: {e}")

async def zero_hunt(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data
    if hunt_data is None: return await update.message.reply_text("⚠️ Upload a file first.")
    
    # Create a temporary DataFrame with numeric goal percentage for filtering
    temp_df = hunt_data.copy()
    temp_df['Goal_Numeric'] = temp_df['Goal Percentage (Hunt)'].str.rstrip('%').astype(float)
    zero = temp_df[(temp_df["Points (Hunt)"]==0)|(temp_df['Goal_Numeric']==0)]
    
    if zero.empty: return await update.message.reply_text("✅ No 0% hunters! Great job!")
    txt=["💀 *0% / 0 Hunt Members*","━━━━━━━━━━━━━━━"]+[f"{i}. {row['Name']} — {row['Goal Percentage (Hunt)']} ({row['Points (Hunt)']} pts)" for i, (_, row) in enumerate(zero.iterrows(), 1)]
    await update.message.reply_text("\n".join(txt), parse_mode="Markdown")

async def clear_data(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data, filtered_players, original_hunt_data
    hunt_data = None
    filtered_players = []
    original_hunt_data = None
    await update.message.reply_text("🧹 Cleared stored hunt data and filtered players list.")

async def search_player(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data
    if hunt_data is None: 
        return await update.message.reply_text("⚠️ Upload a file first.")
    try:
        if not context.args:
            return await update.message.reply_text("Please provide a name to search. \nExample: `/i Player Name`", parse_mode="Markdown")
        search_name = " ".join(context.args)
        results = hunt_data[hunt_data['Name'].str.contains(search_name, case=False, na=False)]
        
        if results.empty:
            return await update.message.reply_text(f"❌ Player not found for: *{search_name}*", parse_mode="Markdown")
        
        if len(results) > 1:
            txt = [f"Found *{len(results)}* matches. Please be more specific:"]
            txt.extend([f"- {name}" for name in results['Name'].head(5)]) 
            if len(results) > 5: txt.append("...and more.")
            return await update.message.reply_text("\n".join(txt), parse_mode="Markdown")

        player = results.iloc[0]
        txt = [
            "🔎 *Player Found*", "━━━━━━━━━━━━━━━",
            f"👤 *Name:* {player['Name']}",
            f"⚔️ *Points:* {player['Points (Hunt)']}",
            f"🎯 *Goal %:* {player['Goal Percentage (Hunt)']}",
            f"📝 *Issues:* {player['Issues'] or 'N/A'}"
        ]
        await update.message.reply_text("\n".join(txt), parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Error during search: {e}")

# --- FIXED: generate_report function ---
async def generate_report(update:Update, context:ContextTypes.DEFAULT_TYPE):
    global hunt_data
    
    if hunt_data is None: 
        return await update.message.reply_text("⚠️ Upload a file first.")
    
    msg = await update.message.reply_text("⏳ Generating updated report...")
    try:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(DOWNLOAD_DIR, f"Updated_Report_{ts}.xlsx")
        
        # Sort by goal percentage before generating report
        temp_df = hunt_data.copy()
        temp_df['Goal_Numeric'] = temp_df['Goal Percentage (Hunt)'].str.rstrip('%').astype(float)
        temp_df = temp_df.sort_values("Goal_Numeric", ascending=False).reset_index(drop=True)
        temp_df = temp_df.drop("Goal_Numeric", axis=1)
        
        hunt_data = temp_df.copy()
        
        hunt_data.to_excel(out_path, index=False)
        format_and_color_excel(out_path)
        
        await msg.edit_text("✅ Updated report generated!")
        await update.message.reply_document(open(out_path,"rb"), filename=os.path.basename(out_path))
    except Exception as e:
        await msg.edit_text(f"❌ Error generating report: {e}")

# ---------- Flask App Setup ----------
app = Flask(__name__)

async def _process_update_async(update_json: dict):
    ptb_app = Application.builder().token(TOKEN).build()

    ptb_app.add_handler(MessageHandler(filters.Document.ALL, handle_file))
    ptb_app.add_handler(CommandHandler("i", search_player))
    ptb_app.add_handler(CommandHandler("summary", summary))
    ptb_app.add_handler(CommandHandler("top10", top10))
    ptb_app.add_handler(CommandHandler("bottom10", bottom10))
    ptb_app.add_handler(CommandHandler("bottom20", bottom20))
    ptb_app.add_handler(CommandHandler("0hunt", zero_hunt))
    ptb_app.add_handler(CommandHandler("clear", clear_data))
    ptb_app.add_handler(CommandHandler("problem", problem))
    ptb_app.add_handler(CommandHandler("generate", generate_report))
    ptb_app.add_handler(CommandHandler("filteredplayers", show_filtered_players))
    ptb_app.add_handler(CommandHandler("include", include_player))
    ptb_app.add_handler(CommandHandler("exclude", exclude_player))
    ptb_app.add_handler(CallbackQueryHandler(handle_player_selection))

    await ptb_app.initialize()
    update = Update.de_json(update_json, ptb_app.bot)
    await ptb_app.process_update(update)
    await ptb_app.update_queue.join()
    await ptb_app.shutdown()

@app.route(f"/{TOKEN}", methods=['POST'])
def webhook():
    update_json = request.get_json()
    asyncio.run(_process_update_async(update_json))
    return 'ok', 200

@app.route("/")
def index():
    return "Bot is running!", 200